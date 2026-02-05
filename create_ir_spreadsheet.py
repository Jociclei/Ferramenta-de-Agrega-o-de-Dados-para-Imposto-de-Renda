from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

wb = Workbook()

# Cores padrão
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
input_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

white_font = Font(bold=True, color='FFFFFF')
black_font = Font(color='000000')
blue_font = Font(color='0000FF')
bold_font = Font(bold=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ABA 1: Dashboard
ws_dashboard = wb.active
ws_dashboard.title = "Dashboard"

ws_dashboard['A1'] = 'AGREGADOR DE DADOS - IMPOSTO DE RENDA'
ws_dashboard['A1'].font = Font(size=16, bold=True, color='366092')
ws_dashboard.merge_cells('A1:E1')
ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws_dashboard['A3'] = f'Ano Base: {datetime.now().year}'
ws_dashboard['A3'].font = bold_font
ws_dashboard['A4'] = f'Última Atualização: {datetime.now().strftime("%d/%m/%Y")}'

ws_dashboard['A6'] = 'NAVEGAÇÃO RÁPIDA'
ws_dashboard['A6'].font = Font(size=14, bold=True, color='366092')

navigation = [
    ['Aba', 'Descrição'],
    ['Rendimentos', 'Salários, pró-labore e outros rendimentos tributáveis'],
    ['Rendimentos Isentos', 'Dividendos, PLR e rendimentos não tributáveis'],
    ['Despesas Dedutíveis', 'Saúde, educação, previdência e dependentes'],
    ['Bens e Direitos', 'Imóveis, veículos, investimentos e outros bens'],
    ['Dívidas', 'Financiamentos, empréstimos e outras dívidas'],
    ['Resumo', 'Consolidação de todos os dados informados']
]

for idx, row in enumerate(navigation, start=7):
    ws_dashboard[f'A{idx}'] = row[0]
    ws_dashboard[f'B{idx}'] = row[1]
    ws_dashboard[f'A{idx}'].font = bold_font if idx == 7 else black_font
    ws_dashboard[f'B{idx}'].font = bold_font if idx == 7 else black_font
    
    if idx == 7:
        ws_dashboard[f'A{idx}'].fill = header_fill
        ws_dashboard[f'B{idx}'].fill = header_fill
        ws_dashboard[f'A{idx}'].font = white_font
        ws_dashboard[f'B{idx}'].font = white_font

ws_dashboard.column_dimensions['A'].width = 20
ws_dashboard.column_dimensions['B'].width = 60

ws_dashboard['A15'] = 'INDICADORES PRINCIPAIS'
ws_dashboard['A15'].font = Font(size=14, bold=True, color='366092')

ws_dashboard['A17'] = 'Total de Rendimentos Tributáveis:'
ws_dashboard['B17'] = '=Rendimentos!B100'
ws_dashboard['B17'].number_format = 'R$ #,##0.00'

ws_dashboard['A18'] = 'Total de Rendimentos Isentos:'
ws_dashboard['B18'] = '=\'Rendimentos Isentos\'!B50'
ws_dashboard['B18'].number_format = 'R$ #,##0.00'

ws_dashboard['A19'] = 'Total de Despesas Dedutíveis:'
ws_dashboard['B19'] = '=\'Despesas Dedutíveis\'!B100'
ws_dashboard['B19'].number_format = 'R$ #,##0.00'

ws_dashboard['A20'] = 'Total de Bens e Direitos:'
ws_dashboard['B20'] = '=\'Bens e Direitos\'!C100'
ws_dashboard['B20'].number_format = 'R$ #,##0.00'

ws_dashboard['A17'].font = bold_font
ws_dashboard['A18'].font = bold_font
ws_dashboard['A19'].font = bold_font
ws_dashboard['A20'].font = bold_font

# ABA 2: Rendimentos
ws_rendimentos = wb.create_sheet("Rendimentos")

ws_rendimentos['A1'] = 'RENDIMENTOS TRIBUTÁVEIS'
ws_rendimentos['A1'].font = Font(size=14, bold=True, color='366092')
ws_rendimentos.merge_cells('A1:F1')

headers = ['Mês', 'Fonte Pagadora', 'CNPJ', 'Tipo de Rendimento', 'Valor Bruto', 'IRRF']
for idx, header in enumerate(headers, start=1):
    cell = ws_rendimentos.cell(row=3, column=idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
         'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

for idx, mes in enumerate(meses, start=4):
    ws_rendimentos[f'A{idx}'] = mes
    ws_rendimentos[f'A{idx}'].fill = input_fill
    ws_rendimentos[f'E{idx}'].number_format = 'R$ #,##0.00'
    ws_rendimentos[f'F{idx}'].number_format = 'R$ #,##0.00'

tipos_rendimento_dv = DataValidation(type="list", formula1='"Salário,Pró-labore,Aluguéis,Outros"', allow_blank=True)
ws_rendimentos.add_data_validation(tipos_rendimento_dv)
tipos_rendimento_dv.add(f'D4:D15')

ws_rendimentos['E100'] = '=SUM(E4:E15)'
ws_rendimentos['F100'] = '=SUM(F4:F15)'
ws_rendimentos['A100'] = 'TOTAL ANUAL'
ws_rendimentos['A100'].font = bold_font
ws_rendimentos['A100'].fill = total_fill
ws_rendimentos['E100'].font = bold_font
ws_rendimentos['E100'].fill = total_fill
ws_rendimentos['F100'].font = bold_font
ws_rendimentos['F100'].fill = total_fill
ws_rendimentos['E100'].number_format = 'R$ #,##0.00'
ws_rendimentos['F100'].number_format = 'R$ #,##0.00'
ws_rendimentos['B100'] = '=SUM(E4:E15)'
ws_rendimentos['B100'].number_format = 'R$ #,##0.00'

ws_rendimentos.column_dimensions['A'].width = 15
ws_rendimentos.column_dimensions['B'].width = 30
ws_rendimentos.column_dimensions['C'].width = 20
ws_rendimentos.column_dimensions['D'].width = 20
ws_rendimentos.column_dimensions['E'].width = 15
ws_rendimentos.column_dimensions['F'].width = 15

# ABA 3: Rendimentos Isentos
ws_isentos = wb.create_sheet("Rendimentos Isentos")

ws_isentos['A1'] = 'RENDIMENTOS ISENTOS E NÃO TRIBUTÁVEIS'
ws_isentos['A1'].font = Font(size=14, bold=True, color='366092')
ws_isentos.merge_cells('A1:D1')

headers_isentos = ['Tipo', 'Fonte Pagadora', 'CNPJ', 'Valor']
for idx, header in enumerate(headers_isentos, start=1):
    cell = ws_isentos.cell(row=3, column=idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

tipos_isentos = ['Dividendos', 'PLR', 'Indenizações', 'Rendimento Poupança', 
                 'Lucro na Venda de Imóvel Residencial', 'Outros']

for idx, tipo in enumerate(tipos_isentos, start=4):
    ws_isentos[f'A{idx}'] = tipo
    ws_isentos[f'A{idx}'].fill = input_fill
    ws_isentos[f'D{idx}'].number_format = 'R$ #,##0.00'

ws_isentos['D50'] = '=SUM(D4:D10)'
ws_isentos['A50'] = 'TOTAL'
ws_isentos['A50'].font = bold_font
ws_isentos['A50'].fill = total_fill
ws_isentos['D50'].font = bold_font
ws_isentos['D50'].fill = total_fill
ws_isentos['D50'].number_format = 'R$ #,##0.00'
ws_isentos['B50'] = '=SUM(D4:D10)'
ws_isentos['B50'].number_format = 'R$ #,##0.00'

ws_isentos.column_dimensions['A'].width = 35
ws_isentos.column_dimensions['B'].width = 30
ws_isentos.column_dimensions['C'].width = 20
ws_isentos.column_dimensions['D'].width = 15

# ABA 4: Despesas Dedutíveis
ws_despesas = wb.create_sheet("Despesas Dedutíveis")

ws_despesas['A1'] = 'DESPESAS DEDUTÍVEIS'
ws_despesas['A1'].font = Font(size=14, bold=True, color='366092')
ws_despesas.merge_cells('A1:E1')

headers_despesas = ['Categoria', 'Beneficiário', 'CPF/CNPJ', 'Descrição', 'Valor']
for idx, header in enumerate(headers_despesas, start=1):
    cell = ws_despesas.cell(row=3, column=idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

ws_despesas['A5'] = 'SAÚDE'
ws_despesas['A5'].font = Font(bold=True, color='FFFFFF')
ws_despesas['A5'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws_despesas.merge_cells('A5:E5')

categorias_saude = ['Plano de Saúde', 'Consultas Médicas', 'Exames', 
                    'Odontologia', 'Fisioterapia', 'Outros - Saúde']
for idx, cat in enumerate(categorias_saude, start=6):
    ws_despesas[f'A{idx}'] = cat
    ws_despesas[f'E{idx}'].number_format = 'R$ #,##0.00'

ws_despesas['E12'] = '=SUM(E6:E11)'
ws_despesas['A12'] = 'Subtotal Saúde'
ws_despesas['A12'].font = bold_font
ws_despesas['E12'].font = bold_font
ws_despesas['E12'].fill = total_fill
ws_despesas['E12'].number_format = 'R$ #,##0.00'

ws_despesas['A14'] = 'EDUCAÇÃO'
ws_despesas['A14'].font = Font(bold=True, color='FFFFFF')
ws_despesas['A14'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_despesas.merge_cells('A14:E14')

categorias_educacao = ['Ensino Fundamental', 'Ensino Médio', 'Ensino Superior', 
                       'Pós-Graduação', 'Ensino Técnico']
for idx, cat in enumerate(categorias_educacao, start=15):
    ws_despesas[f'A{idx}'] = cat
    ws_despesas[f'E{idx}'].number_format = 'R$ #,##0.00'

ws_despesas['E20'] = '=SUM(E15:E19)'
ws_despesas['A20'] = 'Subtotal Educação'
ws_despesas['A20'].font = bold_font
ws_despesas['E20'].font = bold_font
ws_despesas['E20'].fill = total_fill
ws_despesas['E20'].number_format = 'R$ #,##0.00'

ws_despesas['A22'] = 'PREVIDÊNCIA PRIVADA'
ws_despesas['A22'].font = Font(bold=True, color='FFFFFF')
ws_despesas['A22'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ws_despesas.merge_cells('A22:E22')

ws_despesas['A23'] = 'PGBL'
ws_despesas['E23'].number_format = 'R$ #,##0.00'

ws_despesas['A25'] = 'DEPENDENTES'
ws_despesas['A25'].font = Font(bold=True, color='FFFFFF')
ws_despesas['A25'].fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
ws_despesas.merge_cells('A25:E25')

for row in range(26, 30):
    ws_despesas[f'E{row}'].number_format = 'R$ #,##0.00'

ws_despesas['E30'] = '=SUM(E26:E29)'
ws_despesas['A30'] = 'Subtotal Dependentes'
ws_despesas['A30'].font = bold_font
ws_despesas['E30'].font = bold_font
ws_despesas['E30'].fill = total_fill
ws_despesas['E30'].number_format = 'R$ #,##0.00'

ws_despesas['E100'] = '=E12+E20+E23+E30'
ws_despesas['A100'] = 'TOTAL GERAL DEDUTÍVEL'
ws_despesas['A100'].font = Font(bold=True, size=12)
ws_despesas['A100'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws_despesas['A100'].font = Font(bold=True, color='FFFFFF')
ws_despesas['E100'].font = Font(bold=True, size=12)
ws_despesas['E100'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws_despesas['E100'].number_format = 'R$ #,##0.00'
ws_despesas['B100'] = '=E12+E20+E23+E30'
ws_despesas['B100'].number_format = 'R$ #,##0.00'

ws_despesas.column_dimensions['A'].width = 25
ws_despesas.column_dimensions['B'].width = 25
ws_despesas.column_dimensions['C'].width = 18
ws_despesas.column_dimensions['D'].width = 30
ws_despesas.column_dimensions['E'].width = 15

# ABA 5: Bens e Direitos
ws_bens = wb.create_sheet("Bens e Direitos")

ws_bens['A1'] = 'BENS E DIREITOS'
ws_bens['A1'].font = Font(size=14, bold=True, color='366092')
ws_bens.merge_cells('A1:F1')

headers_bens = ['Código', 'Discriminação', 'Localização', 'Valor 31/12/2024', 'Valor 31/12/2025', 'Observações']
for idx, header in enumerate(headers_bens, start=1):
    cell = ws_bens.cell(row=3, column=idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

grupos_bens = [
    ('11', 'Apartamento'),
    ('12', 'Casa'),
    ('13', 'Terreno'),
    ('14', 'Sala ou Conjunto'),
    ('21', 'Veículo Automotor Terrestre'),
    ('31', 'Conta Corrente'),
    ('45', 'Aplicação de Renda Fixa'),
    ('47', 'Fundo de Investimento'),
    ('99', 'Outros Bens e Direitos')
]

row = 4
for codigo, desc in grupos_bens:
    ws_bens[f'A{row}'] = codigo
    ws_bens[f'B{row}'] = desc
    ws_bens[f'D{row}'].number_format = 'R$ #,##0.00'
    ws_bens[f'E{row}'].number_format = 'R$ #,##0.00'
    row += 1

ws_bens['D100'] = '=SUM(D4:D20)'
ws_bens['E100'] = '=SUM(E4:E20)'
ws_bens['A100'] = 'TOTAL'
ws_bens['A100'].font = bold_font
ws_bens['A100'].fill = total_fill
ws_bens['D100'].font = bold_font
ws_bens['D100'].fill = total_fill
ws_bens['E100'].font = bold_font
ws_bens['E100'].fill = total_fill
ws_bens['D100'].number_format = 'R$ #,##0.00'
ws_bens['E100'].number_format = 'R$ #,##0.00'
ws_bens['C100'] = '=E4+E5+E6+E7+E8+E9+E10+E11+E12'
ws_bens['C100'].number_format = 'R$ #,##0.00'

ws_bens.column_dimensions['A'].width = 10
ws_bens.column_dimensions['B'].width = 30
ws_bens.column_dimensions['C'].width = 20
ws_bens.column_dimensions['D'].width = 18
ws_bens.column_dimensions['E'].width = 18
ws_bens.column_dimensions['F'].width = 30

# ABA 6: Dívidas
ws_dividas = wb.create_sheet("Dívidas")

ws_dividas['A1'] = 'DÍVIDAS E ÔNUS REAIS'
ws_dividas['A1'].font = Font(size=14, bold=True, color='366092')
ws_dividas.merge_cells('A1:F1')

headers_dividas = ['Código', 'Discriminação', 'Credor', 'Valor 31/12/2024', 'Valor 31/12/2025', 'Observações']
for idx, header in enumerate(headers_dividas, start=1):
    cell = ws_dividas.cell(row=3, column=idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

tipos_dividas = [
    ('11', 'Financiamento Imobiliário'),
    ('12', 'Financiamento de Veículos'),
    ('13', 'Empréstimo Bancário'),
    ('14', 'Cartão de Crédito'),
    ('99', 'Outras Dívidas')
]

row = 4
for codigo, desc in tipos_dividas:
    ws_dividas[f'A{row}'] = codigo
    ws_dividas[f'B{row}'] = desc
    ws_dividas[f'D{row}'].number_format = 'R$ #,##0.00'
    ws_dividas[f'E{row}'].number_format = 'R$ #,##0.00'
    row += 1

ws_dividas['D50'] = '=SUM(D4:D10)'
ws_dividas['E50'] = '=SUM(E4:E10)'
ws_dividas['A50'] = 'TOTAL'
ws_dividas['A50'].font = bold_font
ws_dividas['A50'].fill = total_fill
ws_dividas['D50'].font = bold_font
ws_dividas['D50'].fill = total_fill
ws_dividas['E50'].font = bold_font
ws_dividas['E50'].fill = total_fill
ws_dividas['D50'].number_format = 'R$ #,##0.00'
ws_dividas['E50'].number_format = 'R$ #,##0.00'

ws_dividas.column_dimensions['A'].width = 10
ws_dividas.column_dimensions['B'].width = 30
ws_dividas.column_dimensions['C'].width = 25
ws_dividas.column_dimensions['D'].width = 18
ws_dividas.column_dimensions['E'].width = 18
ws_dividas.column_dimensions['F'].width = 30

# ABA 7: Resumo
ws_resumo = wb.create_sheet("Resumo")

ws_resumo['A1'] = 'RESUMO CONSOLIDADO - DECLARAÇÃO DE IMPOSTO DE RENDA'
ws_resumo['A1'].font = Font(size=14, bold=True, color='366092')
ws_resumo.merge_cells('A1:C1')

ws_resumo['A3'] = f'Ano-Base: {datetime.now().year}'
ws_resumo['A3'].font = bold_font

ws_resumo['A5'] = 'RENDIMENTOS'
ws_resumo['A5'].font = Font(bold=True, size=12, color='FFFFFF')
ws_resumo['A5'].fill = header_fill
ws_resumo.merge_cells('A5:C5')

dados_resumo = [
    ('Rendimentos Tributáveis', '=Rendimentos!B100'),
    ('(-) IRRF Retido', '=Rendimentos!F100'),
    ('Rendimentos Isentos', '=\'Rendimentos Isentos\'!B50'),
]

row = 6
for desc, formula in dados_resumo:
    ws_resumo[f'A{row}'] = desc
    ws_resumo[f'B{row}'] = formula
    ws_resumo[f'B{row}'].number_format = 'R$ #,##0.00'
    row += 1

ws_resumo['A10'] = 'DEDUÇÕES'
ws_resumo['A10'].font = Font(bold=True, size=12, color='FFFFFF')
ws_resumo['A10'].fill = header_fill
ws_resumo.merge_cells('A10:C10')

ws_resumo['A11'] = 'Total de Despesas Dedutíveis'
ws_resumo['B11'] = '=\'Despesas Dedutíveis\'!B100'
ws_resumo['B11'].number_format = 'R$ #,##0.00'

ws_resumo['A13'] = 'PATRIMÔNIO'
ws_resumo['A13'].font = Font(bold=True, size=12, color='FFFFFF')
ws_resumo['A13'].fill = header_fill
ws_resumo.merge_cells('A13:C13')

patrimonio_data = [
    ('Total de Bens e Direitos (2025)', '=\'Bens e Direitos\'!E100'),
    ('Total de Dívidas (2025)', '=Dívidas!E50'),
    ('Patrimônio Líquido', '=B14-B15'),
]

row = 14
for desc, formula in patrimonio_data:
    ws_resumo[f'A{row}'] = desc
    ws_resumo[f'B{row}'] = formula
    ws_resumo[f'B{row}'].number_format = 'R$ #,##0.00'
    if row == 16:
        ws_resumo[f'A{row}'].font = bold_font
        ws_resumo[f'B{row}'].font = bold_font
        ws_resumo[f'B{row}'].fill = total_fill
    row += 1

ws_resumo['A19'] = 'OBSERVAÇÕES IMPORTANTES'
ws_resumo['A19'].font = Font(bold=True, size=11, color='C65911')
ws_resumo['A19'].fill = warning_fill
ws_resumo.merge_cells('A19:C19')

observacoes = [
    '1. Esta planilha é uma ferramenta auxiliar de organização',
    '2. Sempre confira os dados com documentos oficiais',
    '3. Consulte um contador para orientações específicas',
    '4. Mantenha os comprovantes guardados por 5 anos'
]

row = 20
for obs in observacoes:
    ws_resumo[f'A{row}'] = obs
    ws_resumo.merge_cells(f'A{row}:C{row}')
    row += 1

ws_resumo.column_dimensions['A'].width = 40
ws_resumo.column_dimensions['B'].width = 20
ws_resumo.column_dimensions['C'].width = 20

wb.save('/home/claude/Agregador_IR_2025.xlsx')
print("Planilha criada com sucesso!")
